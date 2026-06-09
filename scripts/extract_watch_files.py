"""One-off: dump the user's watch docs/xlsx to a UTF-8 text file to read."""
import openpyxl, docx2txt, os, io

BASE = r"E:\WATCHES"
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "watch_files_extract.txt")
os.makedirs(os.path.dirname(OUT), exist_ok=True)
buf = io.StringIO()

def dump_xlsx(path):
    buf.write(f"\n{'='*70}\nXLSX: {path}\n{'='*70}\n")
    wb = openpyxl.load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        buf.write(f"\n--- sheet: {ws.title}  ({ws.max_row}x{ws.max_column}) ---\n")
        for row in ws.iter_rows(values_only=True):
            cells = [("" if c is None else str(c)) for c in row]
            # trim trailing empties for readability
            while cells and cells[-1] == "":
                cells.pop()
            if any(cells):
                buf.write(" | ".join(cells) + "\n")

def dump_docx(path):
    buf.write(f"\n{'='*70}\nDOCX: {path}\n{'='*70}\n")
    buf.write(docx2txt.process(path) + "\n")

for f in ["Watch History.xlsx", "Service History.docx", "Watch Full Service Summary.docx"]:
    p = os.path.join(BASE, f)
    (dump_xlsx if f.endswith(".xlsx") else dump_docx)(p)

with open(OUT, "w", encoding="utf-8") as fh:
    fh.write(buf.getvalue())
print(f"wrote {OUT} ({len(buf.getvalue())} chars)")
